from __future__ import annotations

import configparser
import json
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Iterable
from urllib import parse

import certifi
from openpyxl import load_workbook
import requests


SCIM_GROUP_SCHEMA = "urn:ietf:params:scim:schemas:core:2.0:Group"
SCIM_USER_SCHEMA = "urn:ietf:params:scim:schemas:core:2.0:User"
DEFAULT_CUSTOM_ROLE_PREFIX = "PROFILE:t.RGDWWG:"


@dataclass(frozen=True)
class AppConfig:
    tenant_url: str
    token_url: str
    client_id: str
    client_secret: str
    custom_role_prefix: str = DEFAULT_CUSTOM_ROLE_PREFIX
    verify_ssl: bool = True
    timeout_seconds: int = 30
    scim_base_url: str = ""


@dataclass(frozen=True)
class TeamRow:
    team_id: str
    description: str


@dataclass(frozen=True)
class TeamRoleRow:
    team_id: str
    role_id: str


@dataclass(frozen=True)
class UserRow:
    username: str
    team_id: str


@dataclass(frozen=True)
class WorkbookData:
    teams: list[TeamRow]
    team_roles: list[TeamRoleRow]
    users: list[UserRow]


@dataclass(frozen=True)
class CreateTeamOp:
    team_id: str
    description: str


@dataclass(frozen=True)
class AssignRoleOp:
    team_id: str
    role_ids: list[str]


@dataclass(frozen=True)
class AssignUserOp:
    username: str
    team_ids: list[str]


@dataclass(frozen=True)
class ExecutionPlan:
    create_team_ops: list[CreateTeamOp]
    assign_role_ops: list[AssignRoleOp]
    assign_user_ops: list[AssignUserOp]


def workbook_record_counts(data: WorkbookData) -> dict[str, int]:
    return {
        "create_teams": len(data.teams),
        "assign_roles": len(data.team_roles),
        "users": len(data.users),
    }


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_bool(value: str, default: bool = True) -> bool:
    text = _clean(value).lower()
    if text == "":
        return default
    return text in {"1", "true", "yes", "y"}


def load_config(path: Path) -> AppConfig:
    parser = configparser.ConfigParser()
    read_files = parser.read(path, encoding="utf-8")
    if not read_files:
        raise FileNotFoundError(f"Config file not found: {path}")
    if "SAC" not in parser:
        raise ValueError("config.ini must contain a [SAC] section.")

    sac = parser["SAC"]
    required = ["tenant_url", "token_url", "client_id", "client_secret"]
    missing = [field for field in required if _clean(sac.get(field, "")) == ""]
    if missing:
        raise ValueError(f"config.ini is missing required SAC fields: {', '.join(missing)}")

    tenant_url = _clean(sac["tenant_url"]).rstrip("/")
    scim_base_url = _clean(sac.get("scim_base_url")) or f"{tenant_url}/scim2"

    return AppConfig(
        tenant_url=tenant_url,
        token_url=_clean(sac["token_url"]),
        client_id=_clean(sac["client_id"]),
        client_secret=_clean(sac["client_secret"]),
        custom_role_prefix=_clean(sac.get("custom_role_prefix")) or DEFAULT_CUSTOM_ROLE_PREFIX,
        verify_ssl=_to_bool(sac.get("verify_ssl", "true"), default=True),
        timeout_seconds=int(_clean(sac.get("timeout_seconds")) or "30"),
        scim_base_url=scim_base_url.rstrip("/"),
    )


def save_config(config: AppConfig, path: Path) -> None:
    parser = configparser.ConfigParser()
    parser["SAC"] = {
        "tenant_url": config.tenant_url,
        "token_url": config.token_url,
        "client_id": config.client_id,
        "client_secret": config.client_secret,
        "custom_role_prefix": config.custom_role_prefix,
    }
    with path.open("w", encoding="utf-8") as handle:
        parser.write(handle)


def _sheet_rows(path: Path, sheet_name: str):
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Missing worksheet '{sheet_name}' in {path.name}.")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Worksheet '{sheet_name}' is empty.")
    header = [_clean(cell) for cell in rows[0]]
    for raw_row in rows[1:]:
        if all(_clean(cell) == "" for cell in raw_row):
            continue
        yield {header[index]: raw_row[index] if index < len(raw_row) else "" for index in range(len(header))}


def _assert_columns(path: Path, sheet_name: str, required_columns: Iterable[str]) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Missing worksheet '{sheet_name}' in {path.name}.")
    sheet = workbook[sheet_name]
    header_row = next(sheet.iter_rows(values_only=True), None)
    if header_row is None:
        raise ValueError(f"Worksheet '{sheet_name}' is empty.")
    headers = {_clean(cell) for cell in header_row}
    missing = [column for column in required_columns if column not in headers]
    if missing:
        raise ValueError(
            f"Worksheet '{sheet_name}' is missing required columns: {', '.join(missing)}"
        )


def load_workbook_data(path: Path) -> WorkbookData:
    _assert_columns(path, "Create_Teams", ["Team ID", "Team Description"])
    _assert_columns(path, "Assign_Roles", ["TeamID", "RoleID"])
    _assert_columns(path, "Users", ["UserName", "TeamID"])

    teams = [
        TeamRow(team_id=_clean(row["Team ID"]), description=_clean(row.get("Team Description")))
        for row in _sheet_rows(path, "Create_Teams")
        if _clean(row.get("Team ID"))
    ]
    team_roles = [
        TeamRoleRow(team_id=_clean(row["TeamID"]), role_id=_clean(row["RoleID"]))
        for row in _sheet_rows(path, "Assign_Roles")
        if _clean(row.get("TeamID")) and _clean(row.get("RoleID"))
    ]
    users = [
        UserRow(username=_clean(row["UserName"]), team_id=_clean(row["TeamID"]))
        for row in _sheet_rows(path, "Users")
        if _clean(row.get("UserName")) and _clean(row.get("TeamID"))
    ]

    return WorkbookData(teams=teams, team_roles=team_roles, users=users)


def load_workbook_data_from_bytes(payload: bytes) -> WorkbookData:
    with BytesIO(payload) as stream:
        workbook = load_workbook(stream, data_only=True)

    def sheet_rows(sheet_name: str):
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Missing worksheet '{sheet_name}' in uploaded workbook.")
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"Worksheet '{sheet_name}' is empty.")
        header = [_clean(cell) for cell in rows[0]]
        for raw_row in rows[1:]:
            if all(_clean(cell) == "" for cell in raw_row):
                continue
            yield {
                header[index]: raw_row[index] if index < len(raw_row) else ""
                for index in range(len(header))
            }

    def assert_columns(sheet_name: str, required_columns: Iterable[str]) -> None:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Missing worksheet '{sheet_name}' in uploaded workbook.")
        sheet = workbook[sheet_name]
        header_row = next(sheet.iter_rows(values_only=True), None)
        if header_row is None:
            raise ValueError(f"Worksheet '{sheet_name}' is empty.")
        headers = {_clean(cell) for cell in header_row}
        missing = [column for column in required_columns if column not in headers]
        if missing:
            raise ValueError(
                f"Worksheet '{sheet_name}' is missing required columns: {', '.join(missing)}"
            )

    assert_columns("Create_Teams", ["Team ID", "Team Description"])
    assert_columns("Assign_Roles", ["TeamID", "RoleID"])
    assert_columns("Users", ["UserName", "TeamID"])

    teams = [
        TeamRow(team_id=_clean(row["Team ID"]), description=_clean(row.get("Team Description")))
        for row in sheet_rows("Create_Teams")
        if _clean(row.get("Team ID"))
    ]
    team_roles = [
        TeamRoleRow(team_id=_clean(row["TeamID"]), role_id=_clean(row["RoleID"]))
        for row in sheet_rows("Assign_Roles")
        if _clean(row.get("TeamID")) and _clean(row.get("RoleID"))
    ]
    users = [
        UserRow(username=_clean(row["UserName"]), team_id=_clean(row["TeamID"]))
        for row in sheet_rows("Users")
        if _clean(row.get("UserName")) and _clean(row.get("TeamID"))
    ]

    return WorkbookData(teams=teams, team_roles=team_roles, users=users)


def validate_workbook_data(data: WorkbookData) -> list[str]:
    errors: list[str] = []

    seen_teams: set[str] = set()
    for team in data.teams:
        if team.team_id in seen_teams:
            errors.append(f"Create_Teams contains duplicate Team ID '{team.team_id}'.")
        seen_teams.add(team.team_id)

    for role_row in data.team_roles:
        if role_row.team_id == "":
            errors.append("Assign_Roles contains an empty TeamID.")
        if role_row.role_id == "":
            errors.append(f"Assign_Roles contains an empty RoleID for TeamID '{role_row.team_id}'.")

    for user_row in data.users:
        if user_row.username == "":
            errors.append("Users contains an empty UserName.")
        if user_row.team_id == "":
            errors.append(f"Users contains an empty TeamID for user '{user_row.username}'.")

    return errors


def normalize_role_id(role_id: str) -> str:
    return role_id


def build_execution_plan(data: WorkbookData) -> ExecutionPlan:
    create_team_ops = [
        CreateTeamOp(team_id=team.team_id, description=team.description) for team in data.teams
    ]

    grouped_roles: dict[str, list[str]] = {}
    for row in data.team_roles:
        grouped_roles.setdefault(row.team_id, [])
        grouped_roles[row.team_id].append(normalize_role_id(row.role_id))
    assign_role_ops = [
        AssignRoleOp(team_id=team_id, role_ids=role_ids) for team_id, role_ids in grouped_roles.items()
    ]

    grouped_users: dict[str, list[str]] = {}
    for row in data.users:
        grouped_users.setdefault(row.username, [])
        grouped_users[row.username].append(row.team_id)
    assign_user_ops = [
        AssignUserOp(username=username, team_ids=team_ids)
        for username, team_ids in grouped_users.items()
    ]

    return ExecutionPlan(
        create_team_ops=create_team_ops,
        assign_role_ops=assign_role_ops,
        assign_user_ops=assign_user_ops,
    )


def dry_run_report(plan: ExecutionPlan) -> str:
    lines = [
        "DRY RUN: no SAC changes were sent.",
        "",
        f"CREATE TEAMS ({len(plan.create_team_ops)})",
    ]
    for op in plan.create_team_ops:
        lines.append(f"- {op.team_id}: {op.description}")

    lines.append("")
    lines.append(f"ASSIGN ROLES ({len(plan.assign_role_ops)})")
    for op in plan.assign_role_ops:
        lines.append(f"- {op.team_id}: {', '.join(op.role_ids)}")

    lines.append("")
    lines.append(f"ASSIGN USERS ({len(plan.assign_user_ops)})")
    for op in plan.assign_user_ops:
        lines.append(f"- {op.username}: {', '.join(op.team_ids)}")
    return "\n".join(lines)


def _emit_progress(
    progress_callback,
    label: str,
    current: int,
    total: int,
) -> None:
    if progress_callback is not None:
        progress_callback(label, current, total)


class SacScimClient:
    """Our implementation for SAC SCIM provisioning.

    This wraps the existing repo approach, but keeps credentials out of code and
    centralizes the HTTP assumptions in one place.
    """

    def __init__(self, config: AppConfig):
        self.config = config
        self._token: str | None = None
        self._csrf_cache: dict[str, tuple[str | None, str | None]] = {}
        self._active_scim_base_url: str | None = None
        self._trace_log: list[str] = []
        self._session = requests.Session()

    def _trace(self, message: str) -> None:
        self._trace_log.append(message)

    def normalize_role_id(self, role_id: str) -> str:
        cleaned = role_id.strip()
        if cleaned.startswith("PROFILE:"):
            return cleaned
        return f"{self.config.custom_role_prefix}{cleaned}"

    def get_debug_snapshot(self) -> dict[str, object]:
        return {
            "token_url": self.config.token_url,
            "scim_base_url_candidates": self._scim_base_url_candidates(),
            "active_scim_base_url": self._active_scim_base_url or "",
            "trace": list(self._trace_log),
        }

    def _verify_value(self):
        if self.config.verify_ssl:
            return certifi.where()
        return False

    def _is_token_url(self, url: str) -> bool:
        return url.startswith(self.config.token_url)

    def _scim_base_url_candidates(self) -> list[str]:
        configured = self.config.scim_base_url.rstrip("/")
        tenant = self.config.tenant_url.rstrip("/")
        candidates = [configured]

        if configured.endswith("/scim2"):
            candidates.append(f"{tenant}/api/v1/scim")
        elif configured.endswith("/api/v1/scim"):
            candidates.append(f"{tenant}/scim2")
        else:
            candidates.extend([f"{tenant}/scim2", f"{tenant}/api/v1/scim"])

        unique_candidates = []
        seen = set()
        for candidate in candidates:
            if candidate not in seen:
                unique_candidates.append(candidate)
                seen.add(candidate)
        return unique_candidates

    def _current_scim_base_urls(self) -> list[str]:
        if self._active_scim_base_url:
            return [self._active_scim_base_url]
        return self._scim_base_url_candidates()

    def _is_route_not_found_error(self, message: str) -> bool:
        lowered = message.lower()
        return "404" in lowered and "route does not exist" in lowered

    def _request_target_label(self, url: str) -> str:
        if self._is_token_url(url):
            return "token_url"
        if any(url.startswith(base_url) for base_url in self._scim_base_url_candidates()):
            return "SAC SCIM API"
        return url

    def _format_http_error(self, method: str, url: str, status_code: int, body: str) -> str:
        target = self._request_target_label(url)
        body_lower = body.lower()
        if self._is_token_url(url) and status_code in {401, 403}:
            return (
                f"Authentication failed during {method} {target} ({status_code}). "
                "Check client_id, client_secret, OAuth client setup, and User Provisioning access."
            )
        if self._is_token_url(url) and status_code == 404:
            return (
                f"{method} token_url returned 404. Check that the Token URL was copied exactly from the SAC OAuth client details."
            )
        if status_code == 403 and ("csrf" in body_lower or "cookie" in body_lower):
            return (
                f"Possible CSRF or session cookie issue during {method} {target} ({status_code}). "
                f"Response: {body}"
            )
        if status_code in {401, 403}:
            return (
                f"Authorization failed during {method} {target} ({status_code}). "
                f"Response: {body or 'No response body.'} "
                "Check the OAuth client permissions, access token, and SAC-side authorization."
            )
        response_suffix = f" Response: {body}" if body else ""
        return f"HTTP {status_code} during {method} {target}.{response_suffix}"

    def _format_url_error(self, url: str, reason: object) -> str:
        target = self._request_target_label(url)
        reason_text = str(reason)
        if isinstance(reason, requests.exceptions.SSLError) or "CERTIFICATE_VERIFY_FAILED" in reason_text:
            return (
                f"SSL verification failed while connecting to {target}. "
                "Check your local Python certificates, company proxy or SSL inspection, or the endpoint certificate chain."
            )
        return f"Connection failed while calling {target}: {reason_text}"

    def _http(
        self,
        method: str,
        url: str,
        *,
        headers: dict[str, str] | None = None,
        json_body: dict | None = None,
        auth: tuple[str, str] | None = None,
    ) -> tuple[int, dict[str, str], str]:
        try:
            response = self._session.request(
                method=method,
                url=url,
                headers=headers,
                json=json_body,
                auth=auth,
                timeout=self.config.timeout_seconds,
                verify=self._verify_value(),
            )
        except requests.exceptions.RequestException as exc:
            raise RuntimeError(self._format_url_error(url, exc)) from exc

        body = response.text
        if response.status_code >= 400:
            raise RuntimeError(self._format_http_error(method, url, response.status_code, body))
        return response.status_code, dict(response.headers.items()), body

    def get_access_token(self) -> str:
        if self._token:
            self._trace("Reusing cached access token.")
            return self._token

        token_url = self.config.token_url
        separator = "&" if "?" in token_url else "?"
        token_url = f"{token_url}{separator}grant_type=client_credentials"
        self._trace(f"Requesting access token from {token_url}")
        status, _, body = self._http(
            "POST",
            token_url,
            auth=(self.config.client_id, self.config.client_secret),
        )
        if status != 200:
            raise RuntimeError(f"Token request failed with status {status}.")
        parsed = json.loads(body)
        token = parsed.get("access_token")
        if not token:
            raise RuntimeError("Token response did not include access_token.")
        self._token = token
        self._trace("Access token request succeeded.")
        return token

    def _base_headers(self) -> dict[str, str]:
        return {
            "Authorization": f"Bearer {self.get_access_token()}",
            "x-sap-sac-custom-auth": "true",
            "Accept": "application/scim+json",
        }

    def fetch_csrf(self, resource: str) -> tuple[str | None, str | None]:
        if resource in self._csrf_cache:
            self._trace(f"Reusing cached CSRF token for {resource}.")
            return self._csrf_cache[resource]

        last_error: RuntimeError | None = None
        for base_url in self._current_scim_base_urls():
            self._trace(f"Fetching CSRF token for {resource} from {base_url}/{resource}")
            try:
                headers = self._base_headers()
                headers["x-csrf-token"] = "fetch"
                status, response_headers, _ = self._http("GET", f"{base_url}/{resource}", headers=headers)
            except RuntimeError as exc:
                last_error = exc
                self._trace(f"CSRF fetch failed for {base_url}/{resource}: {exc}")
                if self._is_route_not_found_error(str(exc)):
                    continue
                raise
            if status != 200:
                raise RuntimeError(f"Unable to fetch CSRF token for {resource}.")
            csrf = response_headers.get("x-csrf-token")
            cookie = response_headers.get("set-cookie")
            self._active_scim_base_url = base_url
            self._trace(f"CSRF token fetch succeeded for {resource}. Active SCIM base URL: {base_url}")
            self._csrf_cache[resource] = (csrf, cookie)
            return csrf, cookie

        if last_error:
            raise last_error
        raise RuntimeError(f"Unable to fetch CSRF token for {resource}.")

    def _json_request(self, method: str, url: str, payload: dict, csrf_resource: str) -> dict:
        self._trace(f"Sending {method} request to {url}")
        headers = self._base_headers()
        headers["Content-Type"] = "application/json"
        csrf, cookie = self.fetch_csrf(csrf_resource)
        if csrf:
            headers["x-csrf-token"] = csrf
        if cookie:
            headers["Cookie"] = cookie
        _, _, text = self._http(method, url, headers=headers, json_body=payload)
        self._trace(f"{method} request succeeded for {url}")
        return json.loads(text) if text else {}

    def _scim_json_request(self, method: str, path: str, payload: dict, csrf_resource: str) -> dict:
        last_error: RuntimeError | None = None
        for base_url in self._current_scim_base_urls():
            try:
                self._trace(f"Trying SCIM path {base_url}/{path.lstrip('/')}")
                result = self._json_request(method, f"{base_url}/{path.lstrip('/')}", payload, csrf_resource)
                self._active_scim_base_url = base_url
                return result
            except RuntimeError as exc:
                last_error = exc
                self._trace(f"SCIM request failed for {base_url}/{path.lstrip('/')}: {exc}")
                if self._is_route_not_found_error(str(exc)):
                    continue
                raise
        if last_error:
            raise last_error
        raise RuntimeError(f"Unable to call SCIM path: {path}")

    def get_team(self, team_id: str) -> dict | None:
        last_error: RuntimeError | None = None
        for base_url in self._current_scim_base_urls():
            self._trace(f"Checking team at {base_url}/Groups/{parse.quote(team_id)}")
            try:
                _, _, body = self._http(
                    "GET",
                    f"{base_url}/Groups/{parse.quote(team_id)}",
                    headers=self._base_headers(),
                )
                self._active_scim_base_url = base_url
                self._trace(f"Team lookup succeeded for {team_id}. Active SCIM base URL: {base_url}")
                return json.loads(body) if body else None
            except RuntimeError as exc:
                last_error = exc
                self._trace(f"Team lookup failed for {team_id} at {base_url}: {exc}")
                if self._is_route_not_found_error(str(exc)):
                    continue
                if "HTTP 404" in str(exc):
                    return None
                raise
        if last_error:
            raise last_error
        return None

    def get_user(self, username: str) -> dict | None:
        last_error: RuntimeError | None = None
        for base_url in self._current_scim_base_urls():
            self._trace(f"Checking user at {base_url}/Users/{parse.quote(username)}")
            try:
                _, _, body = self._http(
                    "GET",
                    f"{base_url}/Users/{parse.quote(username)}",
                    headers=self._base_headers(),
                )
                self._active_scim_base_url = base_url
                self._trace(f"User lookup succeeded for {username}. Active SCIM base URL: {base_url}")
                return json.loads(body) if body else None
            except RuntimeError as exc:
                last_error = exc
                self._trace(f"User lookup failed for {username} at {base_url}: {exc}")
                if self._is_route_not_found_error(str(exc)):
                    continue
                if "HTTP 404" in str(exc):
                    return None
                raise
        if last_error:
            raise last_error
        return None

    def create_team(self, op: CreateTeamOp) -> None:
        self._trace(f"Step: create team {op.team_id}")
        existing = self.get_team(op.team_id)
        if existing:
            self._trace(f"Team {op.team_id} already exists. Skipping create.")
            return
        payload = {
            "schemas": [SCIM_GROUP_SCHEMA],
            "id": op.team_id,
            "displayName": op.description or op.team_id,
        }
        self._scim_json_request("POST", "Groups", payload, "Groups")

    def assign_roles(self, op: AssignRoleOp) -> None:
        self._trace(f"Step: assign roles to team {op.team_id}")
        team = self.get_team(op.team_id)
        if not team:
            raise RuntimeError(f"Team '{op.team_id}' does not exist.")
        normalized_role_ids = [self.normalize_role_id(role_id) for role_id in op.role_ids]
        self._trace(
            "Normalized role IDs for team "
            f"{op.team_id}: {', '.join(normalized_role_ids)}"
        )
        team["roles"] = normalized_role_ids
        self._scim_json_request(
            "PUT",
            f"Groups/{parse.quote(op.team_id)}",
            team,
            "Groups",
        )

    def assign_user_teams(self, op: AssignUserOp) -> None:
        self._trace(f"Step: assign user {op.username} to teams {', '.join(op.team_ids)}")
        user = self.get_user(op.username)
        if not user:
            raise RuntimeError(
                f"User '{op.username}' does not exist in SAC. "
                "This workbook only provides UserName and TeamID, so the tool can assign existing users only."
            )
        existing_team_ids = {
            entry.get("value") for entry in user.get("groups", []) if entry.get("value")
        }
        user["groups"] = [{"value": team_id} for team_id in sorted(existing_team_ids.union(op.team_ids))]
        self._scim_json_request(
            "PUT",
            f"Users/{parse.quote(op.username)}",
            user,
            "Users",
        )


def run_create_teams(client: SacScimClient, plan: ExecutionPlan, progress_callback=None) -> list[str]:
    lines = []
    total = len(plan.create_team_ops)
    for index, op in enumerate(plan.create_team_ops, start=1):
        client.create_team(op)
        lines.append(f"Team ready: {op.team_id}")
        _emit_progress(progress_callback, "Create Teams", index, total)
    return lines


def run_assign_roles(client: SacScimClient, plan: ExecutionPlan, progress_callback=None) -> list[str]:
    lines = []
    total = len(plan.assign_role_ops)
    for index, op in enumerate(plan.assign_role_ops, start=1):
        client.assign_roles(op)
        lines.append(f"Roles assigned: {op.team_id} -> {', '.join(op.role_ids)}")
        _emit_progress(progress_callback, "Assign Roles", index, total)
    return lines


def run_assign_users(client: SacScimClient, plan: ExecutionPlan, progress_callback=None) -> list[str]:
    lines = []
    total = len(plan.assign_user_ops)
    for index, op in enumerate(plan.assign_user_ops, start=1):
        client.assign_user_teams(op)
        lines.append(f"User assigned: {op.username} -> {', '.join(op.team_ids)}")
        _emit_progress(progress_callback, "Assign Users", index, total)
    return lines
